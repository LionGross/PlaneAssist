# standard library imports
import time
import sys
import argparse
import math
import os
from typing import Dict, Optional

# external library imports
from rich import print
from rich.panel import Panel
from rich.console import Console
from rich.text import Text
from tqdm import tqdm
from openpyxl import load_workbook
from googletrans import Translator
from ambiance import Atmosphere
import requests

disclaimer = """
DISCLAIMER: This Python program is provided for educational purposes only. It calculates the wing area, stall speed,
the thrust required, and other properties of an airplane based on certain input parameters.
The calculations are based on simplified physics and do not take into account all the complex variables and
conditions that can affect an airplane's performance in real-world scenarios.

The program is not intended to be used for actual flight planning or aircraft design. The results should not
be used for making decisions that can result in harm to people or property. The author of this program is
not responsible for any damages or losses arising from the use of this program.

For these calculations the ICAO standard atmosphere 1993 is used.

Additionally, this program includes a language translation feature. The accuracy of translations is not guaranteed,
and the program's author is not responsible for errors or inaccuracies in the translated content.

Always consult with a certified aviation professional or use certified software for flight planning and
aircraft design.

USE AT YOUR OWN RISK.
"""

# default language
language: str = "en"

# initializes global variable for internet connection
internet_connection: bool = False


class PlaneAssist:

    def __init__(self, altitude: float) -> None:
        """
        Initialize the Atmosphere for the given altitude and set the density and gravity attributes.

        :param altitude: Altitude in meters
        :type altitude: float
        """

        atmo = Atmosphere(altitude)
        self.density: float = round(atmo.density[0], 3)
        self.gravity: float = round(atmo.grav_accel[0], 3)

    def menu(self) -> None:
        """
        Displays the PlaneAssist Main Menu and allows the user to select from various options for aircraft calculations.

        :return: None
        """
        while True:
            opt1: str = "[1] - Wing Area Calculator"
            opt2: str = "[2] - Stall Speed Calculator"
            opt3: str = "[3] - Thrust Required Calculator"
            opt4: str = "[4] - Flight Time Calculator"
            opt5: str = "[5] - Range Calculator"
            opt6: str = "[6] - 'All In One' Calculator"
            opt_terminate: str = "[T] - Terminate Program"
            options: str = (
                f"{opt1}\n{opt2}\n{opt3}\n{opt4}\n{opt5}\n{opt6}\n{opt_terminate}"
            )

            translate_n_print("  PlaneAssist Main Menu:", color="cyan")
            translate_n_print(options)

            match input(">>> ").strip():
                case "T":
                    terminate()
                    continue

                case "1":
                    translate_n_print(
                        (
                            "Wing Area Calculator\n"
                            "The following calculations are for horizontal unaccelerated flight."
                        ),
                        color="cyan",
                    )
                    wing_area_prompts = {
                        "cl_max": "Please enter the maximum lift coefficient of your airfoil",
                        "mass": "Please enter the mass of the aircraft (kg)",
                        "velocity": "Please enter the minimum velocity at which your airplane flies (m/s)",
                    }

                    wing_area = self.wing_area_func(manage_data(wing_area_prompts))
                    wing_rec: str = (
                        f"The minimum recommended wing area is {wing_area}m²"
                    )

                    translate_n_print(wing_rec)
                    continue

                case "2":
                    translate_n_print(
                        (
                            "Stall Speed Calculator\n"
                            "The following calculations are for horizontal unaccelerated flight."
                        ),
                        color="cyan",
                    )
                    stall_speed_prompts: Dict[str, str] = {
                        "cl_max": "Please enter the maximum lift coefficient of your airfoil",
                        "mass": "Please enter the mass of the aircraft (kg)",
                        "area": "Please enter the wing area of your airplane (m²)",
                    }

                    stall_speed = self.stall_speed_func(
                        manage_data(stall_speed_prompts)
                    )
                    stall_speed_rec: str = f"The stall speed is {stall_speed}m/s"
                    translate_n_print(stall_speed_rec)
                    continue

                case "3":
                    translate_n_print(
                        (
                            "Thrust Required Calculator\n"
                            "The following calculations are for horizontal unaccelerated flight."
                        ),
                        color="cyan",
                    )
                    thrust_prompts: Dict[str, str] = {
                        "cd": "Please enter the drag coefficient of your airfoil",
                        "velocity": "Please enter the velocity of your airplane (m/s)",
                        "area": "Please enter the wing area of your airplane (m²)",
                    }

                    thrust = self.thrust_func(manage_data(thrust_prompts))
                    thrust_rec: str = f"The minimum thrust required is {thrust}N"
                    translate_n_print(thrust_rec)
                    continue

                case "4":
                    translate_n_print(
                        (
                            "Flight Time Calculator\n"
                            "The following calculations are for horizontal unaccelerated flight."
                        ),
                        color="cyan",
                    )
                    flight_time_prompts: Dict[str, str] = {
                        "capacity": "Please enter the total capacity of your Battery/Battery pack (mAh)",
                        "capacity_used": "Please enter the percentage of your battery you plan to use (%)",
                        "cruise_current_draw": "Please enter the current draw of your plane in cruise (A)",
                        "battery_voltage": "Please enter the battery voltage (V)",
                        "wattage_payload": "Please enter the wattage of all devices "
                        "that are connected to the battery (W)",
                    }

                    flight_time = self.flight_time_func(
                        manage_data(flight_time_prompts)
                    )
                    disclaimer_flight_time: str = (
                        "Note that this calculation does not include "
                        "the increased current draw during takeoff and landing."
                    )
                    flight_time_rec: str = (
                        f"Flight time: {flight_time}minutes\n{disclaimer_flight_time}"
                    )

                    translate_n_print(flight_time_rec)
                    continue

                case "5":
                    translate_n_print(
                        (
                            "Range Calculator\n"
                            "The following calculations are for horizontal unaccelerated flight."
                        ),
                        color="cyan",
                    )
                    range_prompts: Dict[str, str] = {
                        "flight_time": "Please enter the total flight time of your Aircraft (min)",
                        "true_airspeed": "Please enter the true airspeed of your airplane (m/s)",
                        "wind_speed": "Please enter the velocity of the wind (m/s)",
                        "wind_origin": "Please enter the direction from which the wind blows, "
                        "measured clockwise from north (°)",
                        "course": "Please enter the course of your aircraft, measured clockwise from north (°)",
                    }

                    aircraft_range = self.range_func(manage_data(range_prompts))
                    range_rec: str = f"Range: {aircraft_range}km"

                    translate_n_print(range_rec)
                    continue

                case "6":
                    translate_n_print(
                        (
                            "'All In One' Calculator\n"
                            "The following calculations are for horizontal unaccelerated flight."
                        ),
                        color="cyan",
                    )
                    all_in_one_prompts: Dict[str, str] = {
                        "cl_max": "Please enter the maximum lift coefficient of your airfoil",
                        "mass": "Please enter the mass of the aircraft (kg)",
                        "velocity_min": "Please enter the minimum velocity at which your airplane flies (m/s)",
                        "cd": "Please enter the drag coefficient of your airfoil",
                        "capacity": "Please enter the total capacity of your Battery/Battery pack (mAh)",
                        "capacity_used": "Please enter the percentage of your battery you plan to use (%)",
                        "cruise_current_draw": "Please enter the current draw of your plane in cruise (A)",
                        "battery_voltage": "Please enter the battery voltage (V)",
                        "wattage_p": "Please enter the combined wattage of "
                        "all devices that are connected to the battery (W)",
                        "true_airspeed": "Please enter the true airspeed of your airplane (m/s)",
                        "wind_speed": "Please enter the velocity of the wind (m/s)",
                        "wind_origin": "Please enter the direction from which the wind blows, "
                        "measured clockwise from north (°)",
                        "course": "Please enter the course of your aircraft, measured clockwise from north (°)",
                    }

                    wing_area: float
                    stall_speed: float
                    thrust: float
                    flight_time: float
                    aircraft_range: float
                    data: dict

                    input_data = manage_data(all_in_one_prompts)
                    (
                        wing_area,
                        stall_speed,
                        thrust,
                        flight_time,
                        aircraft_range,
                        data,
                    ) = self.all_in_one(input_data)

                    result_wing: str = (
                        f"The minimum recommended wing area is {wing_area}m²"
                    )
                    result_stall: str = f"The stall speed is {stall_speed}m/s"
                    result_thrust: str = f"The minimum thrust required is {thrust}N"
                    result_time: str = f"Flight time: {flight_time}min"
                    result_range: str = f"Range: {aircraft_range}km"
                    all_in_one_rec: str = (
                        f"{result_wing}\n{result_stall}\n"
                        f"{result_thrust}\n{result_time}\n{result_range}"
                    )

                    translate_n_print(all_in_one_rec)
                    translate_n_print(
                        "Do you want to save this data in an .xlsx file?\n"
                        "Press 'y' if yes, else press a random character"
                    )
                    if input(">>> ") == "y":
                        input_data.update(data)
                        save_data(input_data)
                    continue

                case _:
                    translate_n_print(
                        "Sorry, this is an unsupported option.\nPlease try again..."
                    )

    def wing_area_func(self, inputs: Dict[str, float]) -> float:
        """
        Calculate the wing area of the aircraft based on the maximum
        lift coefficient, mass of the aircraft,
        and minimum flying velocity.

        :param inputs: Dictionary containing variable names, used for
            the calculation, as keys and the corresponding value.
        :return: The calculated wing area.
        :rtype: float
        """

        cl_max: float = inputs["cl_max"]
        mass: float = inputs["mass"]
        velocity: float = inputs["velocity"]

        return round(
            ((mass * self.gravity) / (0.5 * self.density * (velocity * 2) * cl_max)),
            2,
        )

    def stall_speed_func(self, inputs: Dict[str, float]) -> float:
        """
        Calculate the stall speed of an aircraft in horizontal
        unaccelerated flight.

        :param inputs: Dictionary containing variable names, used for
            the calculation, as keys and the corresponding value.
        :return: The calculated stall speed of the aircraft.
        :rtype: float
        """

        cl_max: float = inputs["cl_max"]
        mass: float = inputs["mass"]
        area: float = inputs["area"]

        return round(
            math.sqrt((2 * (mass * self.gravity) / self.density * cl_max * area)),
            2,
        )

    def thrust_func(self, inputs: Dict[str, float]) -> float:
        """
        Calculate the thrust required for horizontal unaccelerated flight.

        :param inputs: Dictionary containing variable names, used for
            the calculation, as keys and the corresponding value.
        :return: The calculated thrust required for horizontal
            unaccelerated flight.
        :rtype: float
        """

        cd = inputs["cd"]
        velocity = inputs["velocity"]
        area = inputs["area"]

        return round(
            (0.5 * cd * self.density * velocity**2 * area),
            2,
        )

    @staticmethod
    def flight_time_func(inputs: Dict[str, float]) -> float:
        """
        Calculate the flight time of an electric aircraft for
        horizontal unaccelerated flight.

        :param inputs: Dictionary containing variable names, used for
            the calculation, as keys and the corresponding value.
        :return: The calculated flight time in minutes.
        :rtype: float
        """

        capacity = inputs["capacity"]
        capacity_used = inputs["capacity_used"]
        cruise_current_draw = inputs["cruise_current_draw"]
        wattage_payload = inputs["wattage_payload"]
        battery_voltage = inputs["battery_voltage"]

        return round(
            (
                ((capacity / 1000) * capacity_used * 0.01)
                / (cruise_current_draw + (wattage_payload / battery_voltage))
            )
            * 60,
            2,
        )

    @staticmethod
    def range_func(inputs: Dict[str, float]) -> float:
        """
        Calculate the range of an electric aircraft for
        horizontal unaccelerated flight.

        :param inputs: Dictionary containing variable names, used for
            the calculation, as keys and the corresponding value.
        :return: The calculated range for the given input parameters.
        :rtype: float
        """

        wind_corr_angle = 0
        true_airspeed = inputs["true_airspeed"]
        wind_speed = inputs["wind_speed"]
        course = inputs["course"]
        wind_origin = inputs["wind_origin"]
        flight_time = inputs["flight_time"]

        ground_speed = math.sqrt(
            true_airspeed**2
            + wind_speed**2
            - (
                2 * true_airspeed * wind_speed * math.cos(course)
                - wind_origin
                + wind_corr_angle
            )
        )
        return round(flight_time * 60 * ground_speed / 1000, 2)

    def all_in_one(self, inputs: Dict[str, float]) -> tuple:
        """
        Calculate wing area, stall speed, thrust required, flight time
        and range of an electric aircraft for
        horizontal unaccelerated flight.

        :param inputs: Dictionary containing variable names, used for
            the calculation, as keys and the corresponding value.
        :return: A tuple containing the calculated values for
            wing area, stall speed, thrust, flight time,
            aircraft range, and a dictionary with all the calculated values.
        :rtype: tuple
        """

        wind_corr_angle = 0
        true_airspeed = inputs["true_airspeed"]
        wind_speed = inputs["wind_speed"]
        course = inputs["course"]
        wind_origin = inputs["wind_origin"]
        mass = inputs["mass"]
        cl_max = inputs["cl_max"]
        velocity_min = inputs["velocity_min"]
        cd = inputs["cd"]
        capacity = inputs["capacity"]
        capacity_used = inputs["capacity_used"]
        cruise_current_draw = inputs["cruise_current_draw"]
        battery_voltage = inputs["battery_voltage"]
        wattage_p = inputs["wattage_p"]

        ground_speed = round(
            math.sqrt(
                true_airspeed**2
                + wind_speed**2
                - (
                    2 * true_airspeed * wind_speed * math.cos(course)
                    - wind_origin
                    + wind_corr_angle
                )
            ),
            2,
        )
        wing_area = round(
            (
                (mass * self.gravity)
                / (0.5 * self.density * (velocity_min * 2) * cl_max)
            ),
            2,
        )
        stall_speed = round(
            math.sqrt((2 * (mass * self.gravity) / self.density * cl_max * wing_area)),
            2,
        )
        thrust = round((0.5 * cd * self.density * velocity_min**2 * wing_area))
        flight_time = round(
            (
                ((capacity / 1000) * capacity_used * 0.01)
                / (cruise_current_draw + (wattage_p / battery_voltage))
            )
            * 60,
            2,
        )
        aircraft_range = round(flight_time * 60 * ground_speed / 1000, 2)

        data = {
            "wing_area": wing_area,
            "stall_speed": stall_speed,
            "thrust": thrust,
            "flight_time": flight_time,
            "aircraft_range": aircraft_range,
            "ground_speed": ground_speed,
        }

        return (
            wing_area,
            stall_speed,
            thrust,
            flight_time,
            aircraft_range,
            data,
        )


def main() -> None:
    """
    The main entry point of the PlaneAssist program.

    Clears the terminal screen, checks for validity of
    optional altitude input, initializes the PlaneAssist class,
    checks the internet connection, and displays the welcome message
    and disclaimer. Displays also a progress bar simulation before
    displaying the main menu of the program to encourage reading of
    the disclaimer.

    :return: None
    """
    os.system("clear||cls")
    altitude = arg_checker()
    plane = PlaneAssist(altitude)
    global internet_connection

    if check_internet_connection():
        internet_connection = True

    translate_n_print(
        "Welcome to PlaneAssist. PlaneAssist is a program that helps you "
        "calculate key parameters for aircraft design.\n[i]-- Metric Version[/i]"
    )
    translate_n_print(disclaimer, color="red")

    for _ in tqdm(range(100)):
        time.sleep(0.07)

    plane.menu()


def arg_checker() -> float:
    """
    This function parses the command-line arguments
    to retrieve the altitude and language information.

    :return: Altitude in meters
    :rtype: float
    """
    parser = argparse.ArgumentParser(
        description="This Program calculates certain aircraft parameters"
    )
    parser.add_argument(
        "-a",
        "--altitude",
        type=float,
        nargs="?",
        const=0,
        default=0,
        help="enter height above mean-sea-level as float (use decimal dot)",
    )
    parser.add_argument(
        "-l",
        "--language",
        type=str,
        nargs="?",
        const=0,
        help="enter IETF language tag",
    )
    if parser.parse_args().language:
        global language
        language = parser.parse_args().language

    return parser.parse_args().altitude


def check_internet_connection() -> bool:
    """
    Check for an existing internet connection
    by trying to send a request to google.

    :return: A bool with the state of the internet connection.
    :rtype: bool
    """
    try:
        requests.head("http://www.google.com/", timeout=1)
        return True
    except requests.ConnectionError:
        print(
            Panel(
                "It seems like you are offline.\n"
                "Program will be started using default language (english)..."
            )
        )
        time.sleep(3)
        return False


def translate_n_print(text: str, color: Optional[str] = None) -> None:
    """
    Translates text and prints it on the terminal window.

    Translates the input text to a specified language if an internet
    connection is available and the specified language is not English,
    and then prints the translated text. If there is no internet
    connection the text will be printed as it is (default english).
    If a color is specified, the text is printed in the specified color.

    :param text: Text to be translated and/or printed.
    :param color: Color of the text that is being printed.
    :return: None
    """
    global language, internet_connection
    if internet_connection and language != "en" and language != "english":
        translator = Translator()

        try:
            trans_text: str = translator.translate(text, dest=language).text

        except ValueError:
            language = "en"
            print(
                Panel(
                    "You specified an unsupported language.\nProgram started in default language (english)..."
                )
            )
            time.sleep(3)
            trans_text = translator.translate(text, dest=language).text

    else:
        trans_text: str = text

    if color:
        console = Console()
        styled_text = Text.assemble((trans_text, color))
        console.print(Panel(styled_text))

    else:
        print(Panel(trans_text))


def manage_data(prompt_messages: Dict[str, str]) -> Dict[str, float]:
    """
    Manages the requests for the get_float_input() function,
    stores and organizes the values gathered
    and return a dictionary of input data.

    :param prompt_messages: A dict containing the variables and the
        messages to prompt the user for input.
    :return: dict
    """
    input_data = {}
    for key, message in prompt_messages.items():
        translate_n_print(message)
        input_data[key] = get_float_input()
    return input_data


def get_float_input() -> float:
    """
    Prompts the user to input a float value and handles wrong input.

    :return: The float value entered by the user.
    :rtype: float
    """
    while True:
        try:
            value = float(input(">>> ").strip())
            break

        except ValueError:
            translate_n_print(
                "Input has to be a number.\nAlso make sure to use a decimal dot instead of a comma."
            )
    return value


def save_data(data: dict) -> None:
    """
    Saves the provided data to an Excel file.

    :param data: A dictionary containing the data to be saved.
    :return: None
    """
    wb = load_workbook(filename="DONT_EDIT.xlsx")
    ws = wb.active

    cell_mapping = {
        "cl_max": "C4",
        "cd": "C5",
        "mass": "C7",
        "velocity_min": "C8",
        "capacity": "C10",
        "capacity_used": "C11",
        "cruise_current_draw": "C12",
        "battery_voltage": "C13",
        "wattage_p": "C14",
        "true_airspeed": "C16",
        "wind_speed": "C17",
        "wind_origin": "C18",
        "course": "C19",
        "wing_area": "H5",
        "stall_speed": "H7",
        "thrust": "H9",
        "flight_time": "H11",
        "aircraft_range": "H13",
        "ground_speed": "H16",
    }

    for variable, cell_ref in cell_mapping.items():
        ws[cell_ref] = data[variable]

    translate_n_print(
        "Enter the path where you want to save this file:\n"
        "(if you want the file to be in the local directory just press enter)"
    )
    file_path = input(">>> ").strip()

    while True:
        translate_n_print("enter the name of your file without extension:")
        file_name = input(">>> ").strip()

        if file_name != "":
            break

        else:
            translate_n_print("Name cannot be empty: ")

    wb.save(f"{file_path}{file_name}.xlsx")


def terminate() -> None:
    """
    Prompt the user to confirm if they want to exit the program.
    If confirmed, terminate the program.

    :return: None
    """
    translate_n_print(
        "Are you sure you want to exit? All your progress will be deleted.\n"
        "Press 'y' if yes, else press a random character"
    )
    if input(">>> ") == "y":
        sys.exit("program terminated...")


if __name__ == "__main__":
    main()
